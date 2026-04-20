"""
End-to-end verification of the Excel export in project_revenue.html.

Drives the HTML tool with Playwright, triggers a download, then opens the
resulting .xlsx with openpyxl and checks key cells against the reference
workbook at /Users/lucasknowles/Downloads/Reference Projection Sheet.xlsx.

Run:
  python3 scripts/verify_xlsx_export.py
Requires: playwright (install chromium once: `python3 -m playwright install chromium`)
          openpyxl
"""

import pathlib
import sys
import tempfile

import openpyxl
from playwright.sync_api import sync_playwright

PROJECT = pathlib.Path("/Users/lucasknowles/gw-revenue-projection")
HTML = PROJECT / "project_revenue.html"
REFERENCE = pathlib.Path("/Users/lucasknowles/Downloads/Reference Projection Sheet.xlsx")

# Inputs that match the reference workbook (3-Bedroom, $800/1000/1300 peak ADRs).
FORM_INPUTS = {
    "label": "162 Main Street",
    "bucket": "3br",
    "use_case": "owner_onboarding",
    "peak_central": "1000",
    "peak_conservative": "800",
    "peak_stretch": "1300",
    "turnover_charge": "5500",
    "booking_fee_pct": "16.5",
    "mgmt_rate_pct": "20.0",
}


def fill_and_download(tmp_dir: pathlib.Path) -> pathlib.Path:
    with sync_playwright() as p:
        browser = p.chromium.launch()
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()
        page.goto(HTML.as_uri(), wait_until="networkidle")

        # Wait for ExcelJS to finish loading.
        page.wait_for_function("typeof ExcelJS === 'object'", timeout=5000)

        # Fill text/number inputs; <select> fields need select_option.
        for field in ("label", "peak_central", "peak_conservative", "peak_stretch",
                      "turnover_charge", "booking_fee_pct", "mgmt_rate_pct"):
            page.fill(f"#{field}", FORM_INPUTS[field])
        page.select_option("#bucket", FORM_INPUTS["bucket"])
        page.select_option("#use_case", FORM_INPUTS["use_case"])

        # Run projection.
        page.click("button[type=submit]")
        page.wait_for_selector("#output.visible")

        # Trigger the download.
        with page.expect_download() as dl_info:
            page.click("#download-xlsx")
        dl = dl_info.value
        out_path = tmp_dir / dl.suggested_filename
        dl.save_as(str(out_path))
        browser.close()
        return out_path


def compare(exported: pathlib.Path) -> list[str]:
    """Compare exported workbook to the reference. Return list of failures."""
    failures: list[str] = []

    ref = openpyxl.load_workbook(REFERENCE, data_only=False)
    got = openpyxl.load_workbook(exported, data_only=False)

    # Sheet names may differ (e.g. '3-Bedroom' both sides) but reference uses '3-Bedroom'.
    if got.sheetnames != ref.sheetnames:
        failures.append(f"Sheet names differ: got {got.sheetnames}, ref {ref.sheetnames}")
        return failures

    ws_ref = ref.active
    ws_got = got.active

    # Check a spot-list of cells that matter: titles, headers, formulas, management rate.
    checks = [
        # (cell, expectation_kind, expected_value_or_formula)
        ("A1", "value_contains", "Revenue Projection"),
        ("A2", "value_contains", "Year 3+ Projection"),
        ("D4", "value_eq", "Base ADR"),
        ("F4", "value_eq", "Strong ADR"),
        ("H4", "value_eq", "High ADR"),
        ("C5", "formula_eq", "=31*B5"),
        ("E7", "formula_eq", "=D7*C7"),
        ("E11", "formula_eq", "=SUM(E5:E10)"),
        ("G13", "value_eq", 0.2),                         # 20%
        ("E22", "formula_eq", "=E11"),
        ("G22", "formula_eq", "=G11"),
        ("I22", "formula_eq", "=I11"),
        ("E23", "value_eq", 5500),
        ("E24", "formula_eq", "=E22+E23"),
        ("E27", "formula_eq", "=-E24*0.165"),
        ("E29", "formula_eq", "=-E23"),
        ("E33", "formula_eq", "=E24+E31"),
        ("E36", "formula_eq", "=-(E33)*$G$13"),
        ("E39", "formula_eq", "=E33+E36"),
        ("A20", "value_eq", "Line Item"),
        ("E20", "value_eq", "Base"),
        ("G20", "value_eq", "Strong"),
        ("I20", "value_eq", "High"),
    ]

    for cell_addr, kind, expected in checks:
        actual = ws_got[cell_addr].value
        if kind == "value_eq":
            if actual != expected:
                failures.append(f"{cell_addr}: expected {expected!r}, got {actual!r}")
        elif kind == "value_contains":
            if not (isinstance(actual, str) and expected in actual):
                failures.append(f"{cell_addr}: expected to contain {expected!r}, got {actual!r}")
        elif kind == "formula_eq":
            # openpyxl returns formulas as strings starting with '='.
            if actual != expected:
                failures.append(f"{cell_addr}: expected formula {expected!r}, got {actual!r}")
        else:
            failures.append(f"{cell_addr}: unknown check kind {kind}")

    # Check NET TO HOMEOWNER label cell.
    if ws_got["A39"].value != "NET TO HOMEOWNER":
        failures.append(f"A39: expected 'NET TO HOMEOWNER', got {ws_got['A39'].value!r}")

    # Check fills match on a few landmark cells.
    landmark_fills = [
        ("A4", "FF1F3864"),    # header navy
        ("A21", "FFD9E2F3"),   # section blue
        ("A24", "FFE8EEF7"),   # subtotal
        ("A39", "FFC8E0C8"),   # net green
        ("G13", "FFDAEEF3"),   # editable cyan
    ]
    for cell_addr, want_argb in landmark_fills:
        got_fill = ws_got[cell_addr].fill.fgColor.rgb if ws_got[cell_addr].fill.fgColor else None
        if got_fill != want_argb:
            failures.append(f"{cell_addr} fill: expected {want_argb}, got {got_fill}")

    return failures


def main() -> int:
    if not HTML.exists():
        print(f"HTML not found: {HTML}", file=sys.stderr)
        return 2
    if not REFERENCE.exists():
        print(f"Reference workbook not found: {REFERENCE}", file=sys.stderr)
        return 2

    with tempfile.TemporaryDirectory() as td:
        tmp = pathlib.Path(td)
        print("Driving HTML tool with Playwright…")
        out = fill_and_download(tmp)
        print(f"Downloaded {out.name} ({out.stat().st_size} bytes)")
        failures = compare(out)
        if failures:
            print(f"\n{len(failures)} failure(s):")
            for f in failures:
                print(f"  - {f}")
            return 1
        print("\nAll checks passed.")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
